import getopt
import json
import sys
import datetime
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.chart import PieChart, BubbleChart, Reference, Series
from openpyxl.chart.label import DataLabelList

########## Global variables
disclaimer = """
*** DISCLAIMER ***
This preliminary analysis does NOT guarantee that this application will be approved. 
It should only used as guidance for determining whether an application
should be considered acceptable to START the hardening/approval processes. 
If you are a vendor looking to submit your application, you will be responsible
for rectifying any Failed thresholds prior to onboarding.

Additionally, not all findings may be applicable based on whether or not this 
application must be rebased and/or updated. Yum update or equivalent commands
may rectify a number of these findings but ultimately the application MUST be 
rebased onto an approved base image within Ironbank such as Red Hat UBI8.					
"""

thresholds = {
    'critical':     10,
    'high':         20,
    'medium':       200,
    'low':          -1,
    'negligible':   -1,
    'info':         -1,
    'total':        300
}

# Set the age of CVE thresholds
# The first key (a digit) represents the age of the finding in years. 
# Each year is optional.
# The second set of keys (nested dict) represents the severity thresholds. 
# Each severity is optional. Specify -1 for allow all.
# Example:
#   Current year is 2021.
#   ageThresholds = {
#       0: {                # Would represent year 2021 (current year - 0)
#           'critical': 0,
#           'high':     10,
#           'medium':   20,
#           'low':      -1  # Unlimited low findings
#       },
#       1: {                # Would represent year 2020 (current year - 1)
#           'critical': 3,
#           'high':     5,
#           'medium':   10
#           # low is missing here because it's optional
#       }
#       # Additional years could be specified, but if they're not, they're assumed 
#         to be 0 for all severity levels (except negligible and info).
#   }
ageThresholds = {
    0: {
        'critical': 5,
        'high':     10,
        'medium':   20
    },
    1: {
        'critical': 3,
        'high':     5,
        'medium':   10
    },
    2: {
        'critical': 1,
        'high':     3,
        'medium':   5
    },
    3: {
        'critical': 0,
        'high':     2,
        'medium':   3,
    },
    4: {
        'critical': 0,
        'high':     0,
        'medium':   0
    },
    5: {
        'critical': 0,
        'high':     0,
        'medium':   0
    }
}

######### Sanitize data
def sanitizeData(data):
    df = pd.DataFrame()

    count = 0
    for m in data['matches']:
        # Initialize defaults
        cvss                = ''
        score               = ''
        vector              = ''
        distroType          = ''
        distroVersion       = ''
        links               = ''
        cpes                = ''
        purl                = ''
        metadataName        = ''
        metadataVersion     = ''
        metadataInstalled   = ''
        artifactLocation    = ''

        # Load the appropriate cvss version information (prefer v3 over v2)
        if '3.1' in m['vulnerability']['cvss']['version']:
            cvss = 'cvssV3'
            score = m['vulnerability']['cvss']['metrics']['baseScore']
            vector = m['vulnerability']['cvss']['vector']
        elif '2.0' in m['vulnerability']['cvss']['version']:
            cvss = 'cvssV2'
            score = m['vulnerability']['cvss']['metrics']['baseScore']
            vector = m['vulnerability']['cvss']['vector']

        # Get distribution information
        if 'distro' in m.keys():
            distroType      = m['distro']['name']
            distroVersion   = m['distro']['version']

        if 'urls' in m['vulnerability'].keys():
            links = ', '.join( m['vulnerability']['urls'] )

        for loc in m['artifact']['locations']:
            artifactLocation += loc['path'] + ', '

        # Strip the extra ', ' at the end
        artifactLocation = artifactLocation[:-2]

        if 'metadata' in m['artifact'].keys():
            # Debian packages
            if m['artifact']['metadataType'] == 'DpkgMetadata':
                metadataName        = m['artifact']['metadata']['package']
                metadataVersion     = m['artifact']['metadata']['version']
                for f in m['artifact']['metadata']['files']:
                    metadataInstalled += f['path'] + ', '
                
                # Strip the extra ', ' at the end
                metadataInstalled = metadataInstalled[:-2]

            # Redhat packages
            if m['artifact']['metadataType'] == 'RpmdbMetadata':
                metadataName        = m['artifact']['metadata']['name']
                metadataVersion     = m['artifact']['metadata']['version']
                for f in m['artifact']['metadata']['files']:
                    metadataInstalled += f['path'] + ', '
                
                # Strip the extra ', ' at the end
                metadataInstalled = metadataInstalled[:-2]

            # Python modules
            elif m['artifact']['metadataType'] == 'PythonPackageMetadata':
                metadataName        = m['artifact']['metadata']['name']
                metadataVersion     = m['artifact']['metadata']['version']
                metadataInstalled   = ', '.join( m['artifact']['metadata']
                ['topLevelPackages'] )
            
        # Construct a temporary dataframe
        temp = pd.DataFrame(
            dict(
                cveid                   = m['vulnerability']['id'],
                severity                = m['vulnerability']['severity'],
                cvss                    = cvss,
                score                   = score,
                vector                  = vector,
                links                   = links,
                matcher                 = m['matchDetails']['matcher'],
                distroType              = distroType,
                distroVersion           = distroVersion,
                artifactName            = m['artifact']['name'],
                artifactVersion         = m['artifact']['version'],
                artifactType            = m['artifact']['type'],
                artifactFoundBy         = m['artifact']['foundBy'],
                artifactLocation        = artifactLocation,
                cpes                    = ', '.join( m['artifact']['cpes'] ),
                purl                    = m['artifact']['purl'],
                metadataName            = metadataName,
                metadataVersion         = metadataVersion,
                metadataInstalled       = metadataInstalled,
            ),
            index=[count]
        )

        # Append the temporary dataframe to the actual one
        df = df.append(temp)

        # We use this as the index on the dataframe, so increment it
        count += 1

    return df

######### Get count of findings by severity
def getFindingSeverityCount(df):
    findings = {}
    findings['critical']    = len( df.query('severity == "Critical"') )
    findings['high']        = len( df.query('severity == "High"') )
    findings['medium']      = len( df.query('severity == "Medium"') )
    findings['low']         = len( df.query('severity == "Low"') )
    findings['negligible']  = len( df.query('severity == "Negligible"') )
    findings['info']        = len( df.query('severity == "Info"') )

    return findings

######### Validate thresholds based on severity
def validateSeverityThresholds(df, filename):
    # Get total findings based on criticality levels
    findings = getFindingSeverityCount(df)

    # Get a count of total findings
    findings['total'] = findings['critical'] + findings['high'] + findings['medium'] 
    + findings['low']

    # Print the finding totals (sorted by criticality) to the console
    print("Finding totals:")
    print("%-11s %5s %12s %9s" % ('Severity', 'Count', 'Threshold', 'Status'))

    thresholdsExceeded = 0
    for i in findings:
        status = 'Passed'
        if thresholds[i] == -1:
            threshold = 'None'
        else:
            threshold = thresholds[i]

        if thresholds[i] != -1:
            if findings[i] > thresholds[i]:
                status = 'Failed'
                thresholdsExceeded += 1

        print("%-11s %5d %12s %9s" % (i.capitalize(), findings[i], threshold, status))
    print()

    ##### Print the report to Excel
    # Open previously saved Excel file
    wb = load_workbook(filename)

    # Create severity worksheet
    ws =wb.create_sheet("Severity")
    ws.title = "Severity"

    # Add the severity data
    ws['A1'] = 'Severity'
    ws['B1'] = 'Count'
    ws['C1'] = 'Threshold'
    ws['D1'] = 'Status'

    # Center the severity headers
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['B1'].alignment = Alignment(horizontal='center')
    ws['C1'].alignment = Alignment(horizontal='center')
    ws['D1'].alignment = Alignment(horizontal='center')

    # Bold the severity headers
    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['C1'].font = Font(bold=True)
    ws['D1'].font = Font(bold=True)

    row = 2
    for f in findings:
        ws['A' + str(row)] = f.capitalize()
        ws['B' + str(row)] = findings[f]

        if thresholds[f] != -1:
            ws['C' + str(row)] = thresholds[f]
        else:
            ws['C' + str(row)] = "None"

        if thresholds[f] != -1:
            if findings[f] > thresholds[f]:
                ws['D' + str(row)] = "Failed"
                ws['D' + str(row)].fill = PatternFill(
                    start_color = 'FF0000',
                    end_color = 'FF0000',
                    fill_type = 'solid'
                )
            else:
                ws['D' + str(row)] = "Passed"
                ws['D' + str(row)].fill = PatternFill(
                    start_color = '00FF00',
                    end_color = '00FF00',
                    fill_type = 'solid'
                )
        else:
            ws['D' + str(row)] = "Passed"
            ws['D' + str(row)].fill = PatternFill(
                start_color = '00FF00',
                end_color = '00FF00',
                fill_type = 'solid'
            )

        ws['A' + str(row)].alignment = Alignment(horizontal='center')
        ws['B' + str(row)].alignment = Alignment(horizontal='center')
        ws['C' + str(row)].alignment = Alignment(horizontal='center')
        ws['D' + str(row)].alignment = Alignment(horizontal='center')

        row += 1

    # Create the pie chart
    chart = PieChart()
    labels = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=7)
    data = Reference(ws, min_col=2, min_row=1, max_col=2, max_row=7)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.title="Vulnerabilities"
    chart.height = 12
    chart.width = 15
    ws.add_chart(chart, "F1")

    # Set the disclaimer
    ws.merge_cells('A27:G27')
    ws['A27'] = """*** DISCLAIMER ***
    This preliminary analysis does NOT guarantee that this application will be 
    approved. It should only used as guidance for determining whether an application
    should be considered acceptable to START the hardening/approval processes. 
    If you are a vendor looking to submit your application, you will be responsible
    for rectifying any Failed thresholds prior to onboarding.

    Additionally, not all findings may be applicable based on whether or not this 
    application must be rebased and/or updated. Yum update or equivalent commands
    may rectify a number of these findings but ultimately the application MUST be 
    rebased onto an approved base image within Ironbank such as Red Hat UBI8.
    """
    ws['A27'].alignment = Alignment(wrap_text=True)
    ws.column_dimensions['G'].width = 70
    ws.row_dimensions[27].height = 130

    # Save the Excel document
    wb.save(filename)

    return thresholdsExceeded


########## Verify finding totals by severity and age
def validateAgeThresholds(df, filename):
    thresholdsExceeded = 0

    # Generate age vs severity data (this is estimated by the year listed in the CVE ID)
    adf = pd.DataFrame()
    for i in range(0, 20):
        year = int( datetime.datetime.today().strftime ('%Y') ) - i
        temp = df[ df['cveid'].str.contains(str(year)) ]

        critical    = len( temp.query('severity == "Critical"') )
        high        = len( temp.query('severity == "High"') )
        medium      = len( temp.query('severity == "Medium"') )
        low         = len( temp.query('severity == "Low"') )
        negligible  = len( temp.query('severity == "Negligible"') )
        info        = len( temp.query('severity == "Info"') )


        status = "Passed"

        if i in ageThresholds.keys():
            if 'critical' in ageThresholds[i].keys():
                if critical > ageThresholds[i]['critical']:
                    if ageThresholds[i]['critical'] != -1:
                        status = "Failed"
                        thresholdsExceeded += 1
            if 'high' in ageThresholds[i].keys():
                if high > ageThresholds[i]['high'] and ageThresholds[i]['high'] != -1:
                    status = "Failed"
                    thresholdsExceeded += 1
            if 'medium' in ageThresholds[i].keys():
                if medium > ageThresholds[i]['medium']:
                    if ageThresholds[i]['medium'] != -1:
                        status = "Failed"
                        thresholdsExceeded += 1
            if 'low' in ageThresholds[i].keys():
                if low > ageThresholds[i]['low'] and ageThresholds[i]['low'] != -1:
                    status = "Failed"
                    thresholdsExceeded += 1
            if 'negligible' in ageThresholds[i].keys():
                if negligible > ageThresholds[i]['negligible']:
                    if ageThresholds[i]['negligible'] != -1:
                        status = "Failed"
                        thresholdsExceeded += 1
            if 'info' in ageThresholds[i].keys():
                if info > ageThresholds[i]['info'] and ageThresholds[i]['info'] != -1:
                    status = "Failed"
                    thresholdsExceeded += 1
        else:
            if critical + high + medium > 0:
                status = "Failed"
                thresholdsExceeded += 1
            else:
                status = "Passed"


        tempAdf = pd.DataFrame(
            dict(
                year        = year,
                critical    = critical,
                high        = high,
                medium      = medium,
                low         = low,
                negligible  = negligible,
                info        = info,
                status      = status
            ),
            index=[i]
        )

        adf = adf.append(tempAdf)

    print(adf.to_string(index=False))
    print()

    ##### Print the report to Excel
    # Open previously saved Excel file
    wb = load_workbook(filename)

    # Create age worksheet
    ws =wb.create_sheet("Vulnerabilities by Age")
    ws.title = "Vulnerabilities by Age"

    year = int( datetime.datetime.today().strftime ('%Y') )

    ws['A1'] = "Year"
    ws['B1'] = "Critical"
    ws['C1'] = "High"
    ws['D1'] = "Medium"
    ws['E1'] = "Low"
    ws['F1'] = "Negligible"
    ws['G1'] = "Info"
    ws['H1'] = "Status"

    ws['A1'].alignment = Alignment(horizontal='center')
    ws['B1'].alignment = Alignment(horizontal='center')
    ws['C1'].alignment = Alignment(horizontal='center')
    ws['D1'].alignment = Alignment(horizontal='center')
    ws['E1'].alignment = Alignment(horizontal='center')
    ws['F1'].alignment = Alignment(horizontal='center')
    ws['G1'].alignment = Alignment(horizontal='center')
    ws['H1'].alignment = Alignment(horizontal='center')

    ws['A1'].font = Font(bold=True)
    ws['B1'].font = Font(bold=True)
    ws['C1'].font = Font(bold=True)
    ws['D1'].font = Font(bold=True)
    ws['E1'].font = Font(bold=True)
    ws['F1'].font = Font(bold=True)
    ws['G1'].font = Font(bold=True)
    ws['H1'].font = Font(bold=True)

    rowCount = 2
    for index, row in adf.iterrows():
        ws['A' + str(rowCount)] = row['year']
        ws['B' + str(rowCount)] = row['critical']
        ws['C' + str(rowCount)] = row['high']
        ws['D' + str(rowCount)] = row['medium']
        ws['E' + str(rowCount)] = row['low']
        ws['F' + str(rowCount)] = row['negligible']
        ws['G' + str(rowCount)] = row['info']
        ws['H' + str(rowCount)] = row['status']

        ws['A' + str(rowCount)].alignment = Alignment(horizontal='center')
        ws['B' + str(rowCount)].alignment = Alignment(horizontal='center')
        ws['C' + str(rowCount)].alignment = Alignment(horizontal='center')
        ws['D' + str(rowCount)].alignment = Alignment(horizontal='center')
        ws['E' + str(rowCount)].alignment = Alignment(horizontal='center')
        ws['F' + str(rowCount)].alignment = Alignment(horizontal='center')
        ws['G' + str(rowCount)].alignment = Alignment(horizontal='center')
        ws['H' + str(rowCount)].alignment = Alignment(horizontal='center')

        if row['status'] == 'Passed':
            ws['H' + str(rowCount)].fill = PatternFill(
                start_color = '00FF00',
                end_color = '00FF00',
                fill_type = 'solid'
            )
        else:
            ws['H' + str(rowCount)].fill = PatternFill(
                start_color = 'FF0000',
                end_color = 'FF0000',
                fill_type = 'solid'
            )

        rowCount += 1

    chart = BubbleChart()
    chart.style = 18

    # Criticals
    title = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=2)
    xvalues = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=21)
    yvalues = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=21)
    sizes = Reference(ws, min_col=2, min_row=2, max_col=2, max_row=21)
    series = Series(values=yvalues, xvalues=xvalues, zvalues=sizes, title="Critical")
    series.graphicalProperties.line.solidFill = 'C11F22'
    series.graphicalProperties.solidFill = 'C11F22'
    chart.series.append(series)

    # Highs
    title = Reference(ws, min_col=3, min_row=2, max_col=3, max_row=2)
    xvalues = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=21)
    yvalues = Reference(ws, min_col=3, min_row=2, max_col=3, max_row=21)
    sizes = Reference(ws, min_col=3, min_row=2, max_col=3, max_row=21)
    series = Series(values=yvalues, xvalues=xvalues, zvalues=sizes, title="High")
    series.graphicalProperties.line.solidFill = 'EB8420'
    series.graphicalProperties.solidFill = 'EB8420'
    chart.series.append(series)

    # Mediums
    title = Reference(ws, min_col=4, min_row=2, max_col=4, max_row=2)
    xvalues = Reference(ws, min_col=1, min_row=2, max_col=1, max_row=21)
    yvalues = Reference(ws, min_col=4, min_row=2, max_col=4, max_row=21)
    sizes = Reference(ws, min_col=4, min_row=2, max_col=4, max_row=21)
    series = Series(values=yvalues, xvalues=xvalues, zvalues=sizes, title="Medium")
    series.graphicalProperties.line.solidFill = 'F0D47A'
    series.graphicalProperties.solidFill = 'F0D47A'
    chart.series.append(series)

    int( datetime.datetime.today().strftime ('%Y') )
    chart.x_axis.scaling.min = year - 21
    chart.x_axis.scaling.max = year + 1
    chart.x_axis.scaling.orientation = "maxMin"
    chart.x_axis.majorUnit = 1
    chart.x_axis.minorUnit = 0.2

    chart.dataLabels = DataLabelList() 
    chart.dataLabels.showVal = True

    chart.title="Vulnerabilities by Age"
    chart.width = 40
    chart.height = 23
    ws.add_chart(chart, "J1")

    # Order the worksheets properly
    wb._sheets = [
        wb._sheets[1],  # Summary
        wb._sheets[2],  # Vulnerabilities by age
        wb._sheets[0]   # Findings
    ]

    # Save the Excel document
    wb.save(filename)

    return thresholdsExceeded

########## Format Findings Excel tab
def formatFindings(filename):
    # Open previously saved Excel file
    wb = load_workbook(filename)
    ws = wb['Sheet1']
    ws.title = "Findings"

    # Set column widths
    ws.column_dimensions['B'].width = 19    # CVE ID
    ws.column_dimensions['C'].width = 9     # Severity
    ws.column_dimensions['D'].width = 8     # CVSS version
    ws.column_dimensions['E'].width = 8     # Score
    ws.column_dimensions['F'].width = 37.5  # Vector
    ws.column_dimensions['G'].width = 55    # Links
    ws.column_dimensions['H'].width = 13    # Matcher
    ws.column_dimensions['I'].width = 9     # DistroType
    ws.column_dimensions['J'].width = 10.8  # Distroversion
    ws.column_dimensions['K'].width = 20    # ArtifactName
    ws.column_dimensions['L'].width = 19    # ArtifactVersion
    ws.column_dimensions['M'].width = 9.5   # ArtifactType
    ws.column_dimensions['N'].width = 20    # ArtifactFoundBy
    ws.column_dimensions['O'].width = 50    # ArtifactLocation
    ws.column_dimensions['P'].width = 50    # CPEs
    ws.column_dimensions['Q'].width = 57    # PURL
    ws.column_dimensions['R'].width = 20    # MetadataName
    ws.column_dimensions['S'].width = 19    # MetadataVersion
    ws.column_dimensions['T'].width = 50    # MetadataInstalled

    # Add filtering
    ws.auto_filter.ref = ws.dimensions

    # Freeze top row and first three columns
    ws.freeze_panes = ws['D2']

    # Colorize findings by severity
    for cell in ws['C']:
        
        if cell.value == 'Critical':
            cell.fill = PatternFill(
                start_color = 'FF0000',
                end_color = 'FF0000',
                fill_type = 'solid'
            )
        elif cell.value == 'High':
            cell.fill = PatternFill(
                start_color = 'FF9900',
                end_color = 'FF9900',
                fill_type = 'solid'
            )
        elif cell.value == 'Medium':
            cell.fill = PatternFill(
                start_color = 'FFFF00',
                end_color = 'FF0000',
                fill_type = 'solid'
            )
    
    wb.save(filename)

########## Main function
def main(argv):
    # Process command-line arguments
    filename        = ""
    excelFilename   = ""
    
    usage = []
    usage.append("grype-parser.py <args>")
    usage.append("Argument              Description")
    usage.append("  --filename          Location to the JSON file produced by Grype.")
    usage.append("  --excel-report      Generate an Excel report at this location.")

    try:
        opts, args = getopt.getopt(argv,"h",["filename=","excel-report="])
    except getopt.GetoptError:
        for i in usage:
            print(i)
        sys.exit(2)

    for opt, arg in opts:
        if opt == '-h':
            print(usage)
            sys.exit()
        elif opt in ("--filename"):
            filename = arg
        elif opt in ("--excel-report"):
            excelFilename = arg

    if filename == '':
        print("No filename specified.")
        sys.exit(1)
    if excelFilename == '':
        print("No Excel filename specified.")
        sys.exit(1)


    try:
        f = open(filename)
    except Exception as e:
        print('Unable to open file for reading! ' + str(e))
        sys.exit(1)

    try:    
        data = json.load(f)
    except Exception as e:
        print('Unable to parse JSON file! ' + str(e))
        sys.exit(1)

    f.close()

    # The JSON object is too complex for Pandas, so do some transformations to make it
    # easier to get what we want.
    df = sanitizeData(data)

    # Define the sorting method
    sorting = {
        'Critical':     0,
        'High':         1,
        'Medium':       2,
        'Low':          3,
        'Negligible':   4,
        'Info':         5,
    }

    # Sort the dataframe
    sdf = df.sort_values(by='severity', key=lambda x: x.map(sorting))

    # Save the sorted dataframe to Excel so we can generate some reports later
    sdf.to_excel(excelFilename)
    formatFindings(excelFilename)

    # Generate status reports
    thresholdsExceeded  = validateSeverityThresholds(df, excelFilename)
    thresholdsExceeded += validateAgeThresholds(df, excelFilename)

    # Print final disclaimer about thresholds
    if thresholdsExceeded > 0:
        print("This application exceeds %d thresholds." % (thresholdsExceeded))
    else:
        print("This application Passed all thresholds.")

    print(disclaimer)

    


    

if __name__ == "__main__":
    main(sys.argv[1:])